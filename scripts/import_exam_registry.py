#!/usr/bin/env python3
"""Bulk importer for the exam-registry workbook (State PSC edition).

Usage:
    python scripts/import_exam_registry.py --xlsx PATH [--dry-run] [--verbose]

Sheets processed (in order):
    1. State PSC Detailed Registry  — organizations + exams + cycles
    2. Exam Registry                — exams + cycles (orgs resolved by short-name lookup)

Sheets skipped:
    - PSC Coverage Summary          — aggregate/summary only, no actionable rows
    - Subordinate Boards (Draft)    — DEFERRED (Phase 2 follow-up): rows are self-flagged
                                      as unverified draft; import once data is confirmed.

Sheets processed (in order):
    1. State PSC Detailed Registry  — organizations + exams + cycles
    2. Exam Registry                — exams + cycles (orgs resolved by short-name lookup)
    3. Source URLs                  — central recruiting-body orgs (UPSC/SSC/IBPS/RRB etc.)

Dedupe keys:
    organizations : normalize(short_name).upper() + '|' + (state or '').lower() + '|' + type
    exams         : slugify(state_prefix + '-' + exam_name)
    exam_cycles   : (exam_id, year, cycle_name)  — backed by unique constraint in migration 030

All rows land as metadata.import_status = 'pending_review'.  Nothing is set to
verified or locked.  Operator must review and promote manually.

calendar_status derivation (organizations column, migration 167):
    published    — "Annual Calendar Published?" contains "yes" AND one of:
                   annual calendar / planner / exam calendar / advertisement calendar
    tentative    — "yes" AND one of: tentative / proposed
    partial      — partial / notification-wise / exam schedule / timetable /
                   web notes / monthly programmes
    needs_review — anything else (blank, unclear)

URL policy:
    PSC Source URL     → organizations.metadata.official_url
    Calendar/Schedule URL → exam_cycles.metadata.calendar_url (most-recent cycle only)
    NOT written to source_registry — no consumer needs it; avoids stale-copy burden.
"""
from __future__ import annotations

import argparse
import logging
import os
import re
import sys
from pathlib import Path
from typing import Any

logger = logging.getLogger("import_exam_registry")


# ── path bootstrap ────────────────────────────────────────────────────────────

def _bootstrap_path() -> None:
    repo_root = Path(__file__).resolve().parent.parent
    sys.path.insert(0, str(repo_root / "app" / "backend"))


# ── slug / normalize helpers ──────────────────────────────────────────────────

def slugify(s: str) -> str:
    return re.sub(r"-{2,}", "-",
           re.sub(r"[^a-z0-9]+", "-",
           str(s or "").lower().strip())).strip("-")


def normalize_short_name(raw: str) -> str:
    """Upper-case, strip whitespace, collapse internal spaces."""
    return re.sub(r"\s+", "", str(raw or "").strip().upper())


def org_dedupe_key(short_name: str, state: str | None, org_type: str) -> str:
    return f"{normalize_short_name(short_name)}|{(state or '').lower().strip()}|{org_type}"


def exam_slug(state_prefix: str | None, exam_name: str) -> str:
    prefix = slugify(state_prefix or "national")
    return f"{prefix}-{slugify(exam_name)}"


# ── calendar_status derivation ────────────────────────────────────────────────

_PUBLISHED_KWS = re.compile(
    r"annual\s+calendar|annual\s+planner|exam\s+calendar|advertisement\s+calendar",
    re.I,
)
_TENTATIVE_KWS = re.compile(r"tentative|proposed", re.I)
_PARTIAL_KWS = re.compile(
    r"partial|notification.?wise|exam\s+schedule|exam.?wise\s+calendar"
    r"|timetable|web\s+notes|monthly\s+programme",
    re.I,
)


def derive_calendar_status(raw: str | None) -> str:
    if not raw or not raw.strip():
        return "needs_review"
    s = raw.strip()
    yes_prefix = s.lower().startswith("yes")
    if yes_prefix and _TENTATIVE_KWS.search(s):
        return "tentative"
    if yes_prefix and _PUBLISHED_KWS.search(s):
        return "published"
    if _PARTIAL_KWS.search(s):
        return "partial"
    return "needs_review"


# ── workbook reading ──────────────────────────────────────────────────────────

def _cell(row: dict, *keys: str) -> str | None:
    """Return the first non-blank value from a list of column name candidates."""
    for k in keys:
        v = row.get(k)
        if v is not None and str(v).strip():
            return str(v).strip()
    return None


def load_workbook(path: Path) -> dict[str, list[dict]]:
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl is required: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    result: dict[str, list[dict]] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        # Find header row: first row with at least 2 non-None cells
        header_idx = 0
        for i, row in enumerate(rows):
            non_none = [c for c in row if c is not None]
            if len(non_none) >= 2:
                header_idx = i
                break
        headers = [str(h).strip() if h is not None else f"_col{j}"
                   for j, h in enumerate(rows[header_idx])]
        data = []
        for row in rows[header_idx + 1:]:
            if all(c is None for c in row):
                continue
            data.append(dict(zip(headers, row)))
        result[name] = data
    return result


# ── org upsert ────────────────────────────────────────────────────────────────

def upsert_organization(
    sb: Any,
    *,
    short_name: str,
    full_name: str,
    state: str | None,
    org_type: str,
    calendar_status: str,
    official_url: str | None,
    dry_run: bool,
    org_cache: dict[str, str],
    extra_metadata: dict | None = None,
) -> str | None:
    """Return org id (or None on dry-run). Idempotent by dedupe key.

    extra_metadata: caller-supplied keys merged ON TOP of the standard importer keys
    (import_status, import_source, official_url).  Used by central-body import to add
    import_source='exam_registry_source_urls', source_sheet, source_urls etc.
    Unrelated pre-existing keys are never clobbered (read-modify-write).
    """
    key = org_dedupe_key(short_name, state, org_type)
    if key in org_cache:
        logger.debug("org cache hit: %s", key)
        return org_cache[key]

    def _build_meta(existing: dict | None = None) -> dict:
        base = dict(existing or {})
        base["import_status"] = "pending_review"
        base["import_source"] = "exam_registry_workbook"  # default; extra may override
        if official_url:
            base["official_url"] = official_url
        if extra_metadata:
            base.update(extra_metadata)
        return base

    if not dry_run:
        existing_rows = (
            sb.table("organizations")
            .select("id,name,type,state,calendar_status,metadata")
            .eq("type", org_type)
            .execute()
            .data or []
        )
        for row in existing_rows:
            rkey = org_dedupe_key(
                _abbrev_from_name(row["name"]), row.get("state"), row["type"] or ""
            )
            if rkey == key:
                # Exists — read-modify-write: merge only importer-owned keys so
                # unrelated metadata set by other processes is never clobbered.
                update: dict = {}
                if row.get("calendar_status") != calendar_status:
                    update["calendar_status"] = calendar_status
                existing_meta: dict = row.get("metadata") or {}
                merged_meta = _build_meta(existing_meta)
                if merged_meta != existing_meta:
                    update["metadata"] = merged_meta
                if update:
                    sb.table("organizations").update(update).eq("id", row["id"]).execute()
                org_cache[key] = row["id"]
                logger.debug("org found (existing): %s → %s", key, row["id"])
                return row["id"]

        # Insert new — metadata column added by migration 168
        payload = {
            "name": full_name,
            "type": org_type,
            "state": state,
            "is_active": True,
            "calendar_status": calendar_status,
            "metadata": _build_meta(),
        }
        resp = sb.table("organizations").insert(payload).execute()
        org_id = resp.data[0]["id"]
        org_cache[key] = org_id
        logger.info("org inserted: %s (%s)", full_name, org_id)
        return org_id
    else:
        logger.info("[DRY-RUN] would upsert org: %s | state=%s | type=%s | calendar=%s",
                    full_name, state, org_type, calendar_status)
        org_cache[key] = f"dry-run-{key}"
        return org_cache[key]


def _abbrev_from_name(name: str) -> str:
    """Extract abbreviation: 'Andhra Pradesh PSC' → 'APPSC', 'APPSC' → 'APPSC'."""
    words = name.split()
    # If first word is all-caps abbreviation, use it directly
    if words and re.match(r"^[A-Z]{2,8}$", words[0]):
        return words[0]
    # Build abbreviation from capital letters of content words
    abbrev = "".join(w[0].upper() for w in words if w and w[0].isupper())
    return abbrev or normalize_short_name(name)


# ── exam upsert ───────────────────────────────────────────────────────────────

def upsert_exam(
    sb: Any,
    *,
    slug: str,
    name: str,
    exam_type: str,
    conducting_org_id: str | None,
    dry_run: bool,
    exam_cache: dict[str, str],
) -> str | None:
    if slug in exam_cache:
        return exam_cache[slug]

    if not dry_run:
        existing = sb.table("exams").select("id").eq("slug", slug).execute().data or []
        if existing:
            exam_id = existing[0]["id"]
            # Patch conducting_org_id if missing
            if conducting_org_id:
                sb.table("exams").update(
                    {"conducting_organization_id": conducting_org_id}
                ).eq("id", exam_id).execute()
            exam_cache[slug] = exam_id
            logger.debug("exam found: %s → %s", slug, exam_id)
            return exam_id

        payload = {
            "slug": slug,
            "name": name,
            "exam_type": exam_type,
            "is_active": True,
            "conducting_organization_id": conducting_org_id,
            "metadata": {"import_status": "pending_review"},
        }
        resp = sb.table("exams").insert(payload).execute()
        exam_id = resp.data[0]["id"]
        exam_cache[slug] = exam_id
        logger.info("exam inserted: %s (%s)", name, exam_id)
        return exam_id
    else:
        logger.info("[DRY-RUN] would upsert exam: %s (slug=%s)", name, slug)
        exam_cache[slug] = f"dry-run-{slug}"
        return exam_cache[slug]


# ── cycle upsert ──────────────────────────────────────────────────────────────

def upsert_cycle(
    sb: Any,
    *,
    exam_id: str,
    year: int | None,
    cycle_name: str,
    phases_text: str | None,
    calendar_url: str | None,
    dry_run: bool,
    stats: dict,
) -> None:
    meta: dict = {"import_status": "pending_review"}
    if phases_text:
        meta["typical_phases"] = phases_text
    if calendar_url:
        meta["calendar_url"] = calendar_url

    if dry_run:
        logger.info("[DRY-RUN] would upsert cycle: exam=%s year=%s name=%s",
                    exam_id, year, cycle_name)
        stats["cycles"] += 1
        return

    existing = (
        sb.table("exam_cycles")
        .select("id")
        .eq("exam_id", exam_id)
        .eq("cycle_name", cycle_name)
        .execute()
        .data or []
    )
    if year is not None:
        existing = [
            r for r in (
                sb.table("exam_cycles")
                .select("id,year")
                .eq("exam_id", exam_id)
                .eq("cycle_name", cycle_name)
                .execute()
                .data or []
            )
            if r.get("year") == year
        ]

    if existing:
        sb.table("exam_cycles").update({"metadata": meta}).eq(
            "id", existing[0]["id"]
        ).execute()
        stats["cycles_updated"] = stats.get("cycles_updated", 0) + 1
        logger.debug("cycle updated: %s / %s", exam_id, cycle_name)
    else:
        payload = {
            "exam_id": exam_id,
            "year": year,
            "cycle_name": cycle_name,
            "status": "expected",
            "metadata": meta,
        }
        sb.table("exam_cycles").insert(payload).execute()
        stats["cycles"] += 1
        logger.info("cycle inserted: exam=%s year=%s name=%s", exam_id, year, cycle_name)


# ── sheet processors ──────────────────────────────────────────────────────────

def _parse_year(raw: Any) -> int | None:
    if raw is None:
        return None
    m = re.search(r"(20\d{2})", str(raw))
    return int(m.group(1)) if m else None


def process_state_psc_sheet(
    sb: Any, rows: list[dict], dry_run: bool,
    org_cache: dict, exam_cache: dict, stats: dict
) -> None:
    """State PSC Detailed Registry — 173 rows."""
    for row in rows:
        state = _cell(row, "State/UT", "State")
        short_name = _cell(row, "PSC Short Name", "Short Name") or ""
        full_name = _cell(row, "Conducting Body") or short_name
        exam_family = _cell(row, "Exam/Sub-exam Family", "Exam Family") or ""
        phases_text = _cell(row, "Typical Phases")
        typical_cycle = _cell(row, "Typical Cycle")
        calendar_raw = _cell(row, "Annual Calendar Published?", "Calendar Published?")
        psc_url = _cell(row, "PSC Source URL", "Source URL")
        cal_url = _cell(row, "Calendar/Schedule URL", "Calendar URL")
        exam_type_raw = _cell(row, "Exam Type") or "recruitment"

        if not short_name and not full_name:
            continue

        calendar_status = derive_calendar_status(calendar_raw)
        exam_type = "recruitment" if "recruit" in (exam_type_raw or "").lower() else "recruitment"

        org_id = upsert_organization(
            sb,
            short_name=short_name,
            full_name=full_name,
            state=state,
            org_type="state_psc",
            calendar_status=calendar_status,
            official_url=psc_url,
            dry_run=dry_run,
            org_cache=org_cache,
        )
        stats["orgs"] += 1

        if not exam_family:
            continue

        e_slug = exam_slug(state, exam_family)
        exam_id = upsert_exam(
            sb,
            slug=e_slug,
            name=exam_family,
            exam_type=exam_type,
            conducting_org_id=org_id,
            dry_run=dry_run,
            exam_cache=exam_cache,
        )
        stats["exams"] += 1

        year = _parse_year(typical_cycle)
        cycle_name = typical_cycle or "Annual"
        upsert_cycle(
            sb,
            exam_id=exam_id,
            year=year,
            cycle_name=cycle_name,
            phases_text=phases_text,
            calendar_url=cal_url,
            dry_run=dry_run,
            stats=stats,
        )


def process_exam_registry_sheet(
    sb: Any, rows: list[dict], dry_run: bool,
    org_cache: dict, exam_cache: dict, stats: dict
) -> None:
    """Exam Registry — 225 rows (national-level and state-level exams)."""
    for row in rows:
        exam_name = _cell(row, "Exam") or ""
        conducting_body = _cell(row, "Conducting Body") or ""
        typical_cycle = _cell(row, "Typical Cycle")
        phases_text = _cell(row, "Main Phases")

        if not exam_name:
            continue

        # Resolve org: look up by short-name in cache or by name match
        org_id: str | None = None
        if conducting_body:
            # Try to find from cache using abbreviation heuristic
            abbrev = _abbrev_from_name(conducting_body)
            # Search across all state types
            for key, oid in org_cache.items():
                if key.startswith(normalize_short_name(abbrev) + "|") or \
                   key.startswith(normalize_short_name(conducting_body) + "|"):
                    org_id = oid
                    break
            if not org_id and not dry_run:
                existing = (
                    sb.table("organizations")
                    .select("id,name")
                    .ilike("name", f"%{conducting_body}%")
                    .limit(1)
                    .execute()
                    .data or []
                )
                if existing:
                    org_id = existing[0]["id"]

        # Determine state prefix: if conducting_body looks like a state PSC, extract state
        state_prefix = _extract_state_from_body(conducting_body)
        e_slug = exam_slug(state_prefix, exam_name)
        exam_id = upsert_exam(
            sb,
            slug=e_slug,
            name=exam_name,
            exam_type="recruitment",
            conducting_org_id=org_id,
            dry_run=dry_run,
            exam_cache=exam_cache,
        )
        stats["exams"] += 1

        if typical_cycle:
            year = _parse_year(typical_cycle)
            upsert_cycle(
                sb,
                exam_id=exam_id,
                year=year,
                cycle_name=typical_cycle,
                phases_text=phases_text,
                calendar_url=None,
                dry_run=dry_run,
                stats=stats,
            )


# ── central org / Source URLs sheet ──────────────────────────────────────────

# Canonical name for each central body.  Multiple raw name variants from the
# workbook collapse to a single canonical key so only one org row is created.
# The abbreviation becomes the canonical short_name used for the dedupe key.
_CENTRAL_BODY_ALIASES: dict[str, str] = {
    # UPSC
    "upsc": "UPSC",
    "union public service commission": "UPSC",
    # SSC
    "ssc": "SSC",
    "staff selection commission": "SSC",
    # IBPS
    "ibps": "IBPS",
    "institute of banking personnel selection": "IBPS",
    # RRB / Railway Recruitment Board(s)
    "rrb": "RRB",
    "railway recruitment board": "RRB",
    "railway recruitment boards": "RRB",
    "rrbs": "RRB",
    # RRC (Railway Recruitment Cell)
    "rrc": "RRC",
    "railway recruitment cell": "RRC",
    # SBI / banking
    "sbi": "SBI",
    "state bank of india": "SBI",
    # NTA
    "nta": "NTA",
    "national testing agency": "NTA",
    # DSSSB
    "dsssb": "DSSSB",
    "delhi subordinate services selection board": "DSSSB",
    # LIC
    "lic": "LIC",
    "life insurance corporation": "LIC",
}

_OFFICIAL_SOURCE_TYPES = {"official"}


def _canonical_central_name(raw: str) -> str | None:
    """Return canonical abbreviation for a central body name, or None if unrecognised."""
    key = str(raw or "").lower().strip()
    # Direct lookup
    if key in _CENTRAL_BODY_ALIASES:
        return _CENTRAL_BODY_ALIASES[key]
    # Prefix-match: "RRB Ahmedabad" → "RRB"
    for alias, canonical in _CENTRAL_BODY_ALIASES.items():
        if key.startswith(alias):
            return canonical
    return None


def _group_source_url_rows(rows: list[dict]) -> dict[str, dict]:
    """Group Source URLs sheet rows by canonical body name.

    Returns {canonical_name: {"full_name": str, "urls": [{"type": str, "url": str}]}}
    Rows whose name cannot be mapped to a known central body are counted as skipped.
    """
    groups: dict[str, dict] = {}
    for row in rows:
        # Workbook column name candidates for body name
        raw_name = _cell(row, "Name", "Body", "Organisation", "Organization", "Source") or ""
        raw_type = _cell(row, "Source Type", "Type") or ""
        raw_url = _cell(row, "URL", "Source URL", "Link") or ""

        canonical = _canonical_central_name(raw_name)
        if not canonical:
            continue

        if canonical not in groups:
            groups[canonical] = {"full_name": raw_name, "urls": []}
        if raw_url:
            entry = {"type": raw_type, "url": raw_url}
            if entry not in groups[canonical]["urls"]:
                groups[canonical]["urls"].append(entry)

    return groups


def process_source_urls_sheet(
    sb: Any, rows: list[dict], dry_run: bool,
    org_cache: dict, stats: dict,
) -> None:
    """Source URLs — central recruiting-body orgs (UPSC/SSC/IBPS/RRB etc.)."""
    groups = _group_source_url_rows(rows)
    skipped_ungroupable = len(rows) - sum(len(g["urls"]) for g in groups.values())

    for canonical, group in groups.items():
        full_name = group["full_name"] or canonical
        all_urls = group["urls"]

        # official_url = first URL where Source Type matches "Official"
        official_url: str | None = None
        for entry in all_urls:
            if entry["type"].lower().strip() in _OFFICIAL_SOURCE_TYPES:
                official_url = entry["url"]
                break

        # source_urls: deduped list preserving type/url pairs
        source_urls = all_urls

        calendar_status = "needs_review"  # Source URLs sheet carries no calendar signal

        # Build extra metadata — these are central-body-specific fields merged on top
        # of the standard importer-owned keys set by upsert_organization.
        extra_meta = {
            "import_source": "exam_registry_source_urls",
            "source_sheet": "Source URLs",
            "source_urls": source_urls,
        }

        org_id = upsert_organization(
            sb,
            short_name=canonical,
            full_name=full_name,
            state=None,
            org_type="central_commission",
            calendar_status=calendar_status,
            official_url=official_url,
            dry_run=dry_run,
            org_cache=org_cache,
            extra_metadata=extra_meta,
        )
        stats["central_orgs"] = stats.get("central_orgs", 0) + 1

        if not dry_run:
            logger.info("central org processed: %s (id=%s)", canonical, org_id)
        else:
            logger.info(
                "[DRY-RUN] central org: %s  urls=%d  official_url=%s",
                canonical, len(source_urls), official_url or "(none)",
            )

    if skipped_ungroupable:
        logger.info(
            "Source URLs: %d row(s) skipped (body name not recognised as a central body).",
            skipped_ungroupable,
        )


_STATE_ABBREVS = {
    "andhra pradesh": "andhra-pradesh", "ap": "andhra-pradesh",
    "arunachal pradesh": "arunachal-pradesh",
    "assam": "assam",
    "bihar": "bihar",
    "chhattisgarh": "chhattisgarh",
    "goa": "goa",
    "gujarat": "gujarat",
    "haryana": "haryana",
    "himachal pradesh": "himachal-pradesh", "hp": "himachal-pradesh",
    "jharkhand": "jharkhand",
    "karnataka": "karnataka",
    "kerala": "kerala",
    "madhya pradesh": "madhya-pradesh", "mp": "madhya-pradesh",
    "maharashtra": "maharashtra",
    "manipur": "manipur",
    "meghalaya": "meghalaya",
    "mizoram": "mizoram",
    "nagaland": "nagaland",
    "odisha": "odisha",
    "punjab": "punjab",
    "rajasthan": "rajasthan",
    "sikkim": "sikkim",
    "tamil nadu": "tamil-nadu", "tn": "tamil-nadu",
    "telangana": "telangana",
    "tripura": "tripura",
    "uttar pradesh": "uttar-pradesh", "up": "uttar-pradesh",
    "uttarakhand": "uttarakhand",
    "west bengal": "west-bengal", "wb": "west-bengal",
    "delhi": "delhi",
    "jammu and kashmir": "jammu-kashmir", "j&k": "jammu-kashmir",
    "ladakh": "ladakh",
}


def _extract_state_from_body(body: str) -> str | None:
    b = body.lower()
    for state, slug_prefix in _STATE_ABBREVS.items():
        if state in b:
            return slug_prefix
    return None


# ── main ──────────────────────────────────────────────────────────────────────

def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Import exam registry workbook.")
    parser.add_argument("--xlsx", required=True, type=Path, help="Path to .xlsx workbook.")
    parser.add_argument("--dry-run", action="store_true", help="Preview only; no DB writes.")
    parser.add_argument("--verbose", "-v", action="store_true")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
    )

    _bootstrap_path()

    if args.dry_run:
        logger.info("=== DRY-RUN MODE — no rows will be written ===")
        sb = None
    else:
        from app.db.supabase_client import get_supabase_admin
        sb = get_supabase_admin()

    sheets = load_workbook(args.xlsx)

    # Sanity-check expected sheets
    found = set(sheets.keys())
    logger.info("Sheets found: %s", sorted(found))

    # Subordinate Boards — always count and report; never silently drop
    sub_board_rows = len(sheets.get("Subordinate Boards (Draft)", []))
    if "Subordinate Boards (Draft)" in found:
        logger.info(
            "SKIP: 'Subordinate Boards (Draft)' — DEFERRED to Phase 2 follow-up. "
            "%d rows; self-flagged draft/unverified. Import once data is confirmed.",
            sub_board_rows,
        )

    if "PSC Coverage Summary" in found:
        logger.info("SKIP: 'PSC Coverage Summary' — aggregate only, no actionable rows.")

    org_cache: dict[str, str] = {}
    exam_cache: dict[str, str] = {}
    stats: dict[str, int] = {"orgs": 0, "exams": 0, "cycles": 0, "cycles_updated": 0,
                              "central_orgs": 0}

    # Process sheets: PSC first (populates org_cache), then Exam Registry, then central.
    for sheet_name, processor in [
        ("State PSC Detailed Registry", process_state_psc_sheet),
        ("Exam Registry", process_exam_registry_sheet),
    ]:
        if sheet_name not in sheets:
            logger.warning("Sheet not found: %s", sheet_name)
            continue
        logger.info("Processing sheet: %s (%d rows)", sheet_name, len(sheets[sheet_name]))
        processor(sb, sheets[sheet_name], args.dry_run, org_cache, exam_cache, stats)

    # Process Source URLs sheet — central recruiting-body orgs
    source_url_rows = sheets.get("Source URLs", [])
    if source_url_rows:
        logger.info("Processing sheet: Source URLs (%d rows)", len(source_url_rows))
        process_source_urls_sheet(sb, source_url_rows, args.dry_run, org_cache, stats)
    elif "Source URLs" not in found:
        logger.info("Source URLs sheet not present in workbook.")

    # Report undated-phase backlog count
    undated_phases = 0
    if not args.dry_run and sb:
        result = (
            sb.table("exam_phases")
            .select("id", count="exact")
            .is_("phase_start", "null")
            .execute()
        )
        undated_phases = getattr(result, "count", None) or 0

    logger.info(
        "Import complete. state_orgs=%d  central_orgs=%d  exams=%d  "
        "cycles_inserted=%d  cycles_updated=%d",
        stats["orgs"], stats.get("central_orgs", 0),
        stats["exams"], stats["cycles"], stats.get("cycles_updated", 0),
    )
    if not args.dry_run:
        logger.info("Undated exam_phases backlog (phase_start IS NULL): %d", undated_phases)

    return 0


if __name__ == "__main__":
    sys.exit(main())
